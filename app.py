from flask import Flask, render_template, request, jsonify, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import os
from sqlalchemy import func, extract
from sqlalchemy.sql import case
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
import openpyxl
try:
    import google.generativeai as genai
    HAS_GENAI = True
except ImportError:
    HAS_GENAI = False
    logging.warning("google.generativeai module not found. AI features will be disabled.")
from io import BytesIO
import json
import logging
import re
from bs4 import BeautifulSoup

def parse_turkish_amount(amount_str):
    if not amount_str:
        return 0.0
    # Remove dots (thousands separator) and replace comma with dot (decimal separator)
    clean_str = amount_str.replace('.', '').replace(',', '.')
    # Remove any non-numeric chars except the decimal dot and minus sign
    clean_str = re.sub(r'[^\d.-]', '', clean_str)
    try:
        return float(clean_str)
    except ValueError:
        return 0.0

app = Flask(__name__)
app.config['SECRET_KEY'] = 'a-very-secret-key-that-you-should-change'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///dashboard.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password_hash = db.Column(db.String(150), nullable=False)
    currency = db.Column(db.String(5), nullable=False, default='$')
    gemini_api_key = db.Column(db.String(255), nullable=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Account(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    type = db.Column(db.String(50), nullable=False, default='checking')
    balance = db.Column(db.Float, nullable=False, default=0.0)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    transactions = db.relationship('Transaction', backref='account', cascade="all, delete-orphan", lazy=True)

class Category(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    type = db.Column(db.String(10), nullable=False) # 'income' or 'expense'
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    transactions = db.relationship('Transaction', backref='category', lazy=True)

class Transaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(100), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    account_id = db.Column(db.Integer, db.ForeignKey('account.id'), nullable=False)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=True)

class RecurringTransaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    frequency = db.Column(db.String(10), nullable=False)
    start_date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    next_due_date = db.Column(db.DateTime, nullable=False)
    end_date = db.Column(db.DateTime)
    account_id = db.Column(db.Integer, db.ForeignKey('account.id'), nullable=False)
    account = db.relationship('Account', backref=db.backref('recurring_transactions', cascade='all, delete-orphan'))
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    is_active = db.Column(db.Boolean, default=True, nullable=False)


with app.app_context():
    db.create_all()

def calculate_next_due(start_date, frequency):
    if frequency == 'daily':
        return start_date + timedelta(days=1)
    elif frequency == 'weekly':
        return start_date + timedelta(weeks=1)
    elif frequency == 'monthly':
        return start_date + relativedelta(months=1)
    elif frequency == 'yearly':
        return start_date + relativedelta(years=1)
    return None

def process_recurring_transactions(user_id):
    today = datetime.utcnow().date()
    due_items = RecurringTransaction.query.filter(
        RecurringTransaction.user_id == user_id,
        RecurringTransaction.is_active == True,
        db.func.date(RecurringTransaction.next_due_date) <= today
    ).all()

    for rt in due_items:
        current_due_date = rt.next_due_date
        
        while current_due_date.date() <= today:
            if rt.end_date and current_due_date.date() > rt.end_date.date():
                rt.is_active = False
                break

            new_trans = Transaction(
                description=rt.name,
                amount=rt.amount,
                date=current_due_date,
                account_id=rt.account_id,
                category_id=rt.category_id
            )
            db.session.add(new_trans)
            
            if rt.account:
                rt.account.balance += rt.amount
            
            next_due = calculate_next_due(current_due_date, rt.frequency)
            if not next_due:
                rt.is_active = False
                break
            
            current_due_date = next_due

        rt.next_due_date = current_due_date
        
        if rt.end_date and rt.next_due_date.date() > rt.end_date.date():
            rt.is_active = False

    db.session.commit()

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

@app.template_filter()
def format_number(value):
    if value is None:
        value = 0.0
    try:
        value = float(value)
        s = '{:,.2f}'.format(value)
        return s.replace(',', 'X').replace('.', ',').replace('X', '.')
    except (ValueError, TypeError):
        if isinstance(value, str):
            return value
        return '0,00'
app.jinja_env.filters['format_number'] = format_number

@app.before_request
def check_for_setup():
    if not User.query.first() and request.endpoint not in ['setup', 'static', 'account_details']:
        return redirect(url_for('setup'))
    if User.query.first() and request.endpoint == 'setup':
        return redirect(url_for('login'))

@app.route('/setup', methods=['GET', 'POST'])
def setup():
    if User.query.first():
        return redirect(url_for('login'))
        
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if not username or not password:
            flash('Username and password are required.', 'error')
            return redirect(url_for('setup'))
        
        new_user = User(username=username)
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()
        
        flash('Admin account created successfully! Please log in.', 'success')
        return redirect(url_for('login'))

    return render_template('setup.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()

        if user and user.check_password(password):
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password.', 'error')

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/account/<int:account_id>')
@login_required
def account_details(account_id):
    account = Account.query.filter_by(id=account_id, user_id=current_user.id).first_or_404()
    
    transactions = db.session.query(Transaction, Category.name.label('category_name'))\
        .outerjoin(Category, Transaction.category_id == Category.id)\
        .filter(Transaction.account_id == account.id)\
        .order_by(Transaction.date.desc())\
        .all()

    today = datetime.utcnow()
    six_months_ago = today - timedelta(days=180)

    chart_data_query = db.session.query(
        extract('year', Transaction.date).label('year'),
        extract('month', Transaction.date).label('month'),
        func.sum(case((Transaction.amount > 0, Transaction.amount), else_=0)).label('income'),
        func.sum(case((Transaction.amount < 0, Transaction.amount), else_=0)).label('expense')) \
        .filter(Transaction.account_id == account_id) \
        .filter(Transaction.date >= six_months_ago) \
        .group_by('year', 'month') \
        .order_by('year', 'month') \
        .all()

    labels = []
    income_data = []
    expense_data = []
    for row in chart_data_query:
        month_name = datetime(row.year, row.month, 1).strftime('%b \'%y')
        labels.append(month_name)
        income_data.append(row.income or 0)
        expense_data.append(abs(row.expense or 0))

    chart_json = { 'labels': labels, 'income': income_data, 'expense': expense_data }

    total_income = sum(income_data)
    total_expense = sum(expense_data)
    net_flow = total_income - total_expense
    summary_stats = {
        'total_income': total_income,
        'total_expense': total_expense,
        'net_flow': net_flow
    }

    return render_template(
        'account_details.html', 
        account=account, 
        transactions=transactions,
        chart_data=chart_json,
        summary_stats=summary_stats
    )


@app.route('/api/dashboard', methods=['GET'])
@login_required
def get_dashboard_data():
    process_recurring_transactions(current_user.id)
    accounts = Account.query.filter_by(user_id=current_user.id).order_by(Account.name).all()
    user_account_ids = [acc.id for acc in accounts]

    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    
    monthly_stats = db.session.query(
        func.sum(case((Transaction.amount > 0, Transaction.amount), else_=0)).label('income'),
        func.sum(case((Transaction.amount < 0, Transaction.amount), else_=0)).label('expense')
    ).filter(Transaction.account_id.in_(user_account_ids)) \
     .filter(Transaction.date >= thirty_days_ago) \
     .first()

    total_balance = sum(acc.balance for acc in accounts)
    total_income = monthly_stats.income or 0.0
    total_expense = abs(monthly_stats.expense or 0.0)

    expense_breakdown_data = db.session.query(
        Category.name,
        func.sum(Transaction.amount).label('total')
    ).outerjoin(Category, Transaction.category_id == Category.id) \
     .filter(Transaction.account_id.in_(user_account_ids)) \
     .filter(Transaction.amount < 0) \
     .filter(Transaction.date >= thirty_days_ago) \
     .group_by(Category.name) \
     .order_by(func.sum(Transaction.amount).asc()) \
     .all()

    recent_transactions_data = db.session.query(Transaction, Category.name.label('category_name'))\
        .outerjoin(Category, Transaction.category_id == Category.id)\
        .filter(Transaction.account_id.in_(user_account_ids))\
        .order_by(Transaction.date.desc())\
        .limit(10) \
        .all()
    
    accounts_data = []
    for account in accounts:
        transactions = db.session.query(Transaction, Category.name.label('category_name'))\
            .outerjoin(Category, Transaction.category_id == Category.id)\
            .filter(Transaction.account_id == account.id)\
            .order_by(Transaction.date.desc())\
            .limit(5)\
            .all()
        
        accounts_data.append({
            'id': account.id,
            'name': account.name,
            'type': account.type,
            'balance': account.balance,
            'transactions': [{
                'id': t.Transaction.id, 
                'description': t.Transaction.description, 
                'amount': t.Transaction.amount, 
                'date': t.Transaction.date.isoformat(), 
                'category_id': t.Transaction.category_id,
                'category': t.category_name or 'Uncategorized'
            } for t in transactions]
        })

    return jsonify({
        'summaryStats': {
            'totalBalance': total_balance,
            'totalIncome': total_income,
            'totalExpense': total_expense
        },
        'expenseBreakdown': [{'category': r.name or 'Uncategorized', 'total': abs(r.total or 0)} for r in expense_breakdown_data],
        'recentTransactions': [{
            'id': t.Transaction.id, 
            'description': t.Transaction.description, 
            'amount': t.Transaction.amount, 
            'date': t.Transaction.date.isoformat(), 
            'category': t.category_name or 'Uncategorized'
        } for t in recent_transactions_data],
        'accounts': accounts_data
    })


@app.route('/api/accounts', methods=['POST'])
@login_required
def add_account():
    data = request.get_json()
    name = data.get('name', '').strip()
    account_type = data.get('type', 'checking').strip()
    balance = data.get('balance', 0.0)
    if not name:
        return jsonify({'error': 'Account name is required'}), 400

    new_account = Account(name=name, type=account_type, balance=balance, user_id=current_user.id)
    db.session.add(new_account)
    db.session.commit()

    return jsonify({'id': new_account.id, 'name': new_account.name, 'type': new_account.type, 'balance': new_account.balance, 'transactions': []}), 201

@app.route('/api/accounts/<int:account_id>', methods=['PUT'])
@login_required
def update_account(account_id):
    account = Account.query.filter_by(id=account_id, user_id=current_user.id).first_or_404()
    data = request.get_json()
    name = data.get('name', '').strip()
    account_type = data.get('type', 'checking').strip()
    if not name:
        return jsonify({'error': 'Account name is required'}), 400
    
    account.name = name
    account.type = account_type
    db.session.commit()
    
    return jsonify({'id': account.id, 'name': account.name, 'type': account.type})

@app.route('/api/accounts/<int:account_id>', methods=['DELETE'])
@login_required
def delete_account(account_id):
    account = Account.query.filter_by(id=account_id, user_id=current_user.id).first_or_404()
    db.session.delete(account)
    db.session.commit()
    return jsonify({'message': 'Account deleted successfully'})

@app.route('/api/transactions', methods=['POST'])
@login_required
def add_transaction():
    data = request.get_json()
    account_id = data.get('account_id')
    description = data.get('description', '').strip()
    amount = data.get('amount')
    category_id = data.get('category_id')

    if not all([account_id, description, amount is not None]):
        return jsonify({'error': 'Missing data'}), 400

    account = Account.query.filter_by(id=account_id, user_id=current_user.id).first_or_404()
    
    new_transaction = Transaction(description=description, amount=amount, category_id=category_id, account_id=account.id)
    
    account.balance += new_transaction.amount
    
    db.session.add(new_transaction)
    db.session.commit()

    return jsonify({
        'id': new_transaction.id, 
        'description': new_transaction.description, 
        'amount': new_transaction.amount,
        'date': new_transaction.date.isoformat(),
        'category_id': new_transaction.category_id,
        'new_balance': account.balance
    }), 201

@app.route('/api/transactions/<int:transaction_id>', methods=['PUT'])
@login_required
def update_transaction(transaction_id):
    transaction = Transaction.query.get_or_404(transaction_id)
    if transaction.account.user_id != current_user.id:
        return jsonify({'error': 'Forbidden'}), 403

    data = request.get_json()
    description = data.get('description', '').strip()
    amount = data.get('amount')
    category_id = data.get('category_id')

    if not description or amount is None:
        return jsonify({'error': 'Description and amount are required'}), 400
    
    old_amount = transaction.amount
    transaction.description = description
    transaction.amount = amount
    transaction.category_id = category_id

    transaction.account.balance = transaction.account.balance - old_amount + amount
    
    db.session.commit()
    return jsonify({
        'id': transaction.id, 
        'description': transaction.description, 
        'amount': transaction.amount,
        'category_id': transaction.category_id
    })

@app.route('/api/transactions/<int:transaction_id>', methods=['DELETE'])
@login_required
def delete_transaction(transaction_id):
    transaction = Transaction.query.get_or_404(transaction_id)
    if transaction.account.user_id != current_user.id:
        return jsonify({'error': 'Forbidden'}), 403

    transaction.account.balance -= transaction.amount
    
    db.session.delete(transaction)
    db.session.commit()
    return jsonify({'message': 'Transaction deleted successfully'})

@app.route('/api/categories', methods=['GET'])
@login_required
def get_categories():
    categories = Category.query.filter_by(user_id=current_user.id).order_by(Category.type, Category.name).all()
    return jsonify([{'id': c.id, 'name': c.name, 'type': c.type} for c in categories])

@app.route('/api/categories', methods=['POST'])
@login_required
def add_category():
    data = request.get_json()
    name = data.get('name', '').strip()
    category_type = data.get('type', '').strip()
    if not name or not category_type:
        return jsonify({'error': 'Name and type are required'}), 400
    if category_type not in ['income', 'expense']:
        return jsonify({'error': 'Invalid category type'}), 400

    new_category = Category(name=name, type=category_type, user_id=current_user.id)
    db.session.add(new_category)
    db.session.commit()
    return jsonify({'id': new_category.id, 'name': new_category.name, 'type': new_category.type}), 201

@app.route('/api/categories/<int:category_id>', methods=['DELETE'])
@login_required
def delete_category(category_id):
    category = Category.query.filter_by(id=category_id, user_id=current_user.id).first_or_404()
    db.session.delete(category)
    db.session.commit()
    return jsonify({'message': 'Category deleted successfully'})

@app.route('/api/settings', methods=['GET', 'PUT'])
@login_required
def user_settings():
    if request.method == 'PUT':
        data = request.get_json()
        currency = data.get('currency')
        api_key = data.get('gemini_api_key')
        
        if currency and len(currency) <= 5:
            current_user.currency = currency
            
        if api_key is not None: # check for None to distinguish from empty string
            current_user.gemini_api_key = api_key.strip()
            
        db.session.commit()
        return jsonify({'message': 'Settings updated successfully'})
    
    return jsonify({
        'username': current_user.username,
        'currency': current_user.currency,
        'gemini_api_key': current_user.gemini_api_key
    })

@app.route('/api/recurring', methods=['GET'])
@login_required
def get_recurring_transactions():
    recurring_txns = RecurringTransaction.query.filter_by(user_id=current_user.id).order_by(RecurringTransaction.next_due_date.asc()).all()
    results = []
    for r in recurring_txns:
        results.append({
            'id': r.id,
            'name': r.name,
            'amount': r.amount,
            'frequency': r.frequency,
            'start_date': r.start_date.isoformat(),
            'next_due_date': r.next_due_date.isoformat(),
            'end_date': r.end_date.isoformat() if r.end_date else None,
            'account_id': r.account_id,
            'account_name': r.account.name if r.account else 'N/A',
            'category_id': r.category_id,
            'is_active': r.is_active
        })
    return jsonify(results)

@app.route('/api/recurring', methods=['POST'])
@login_required
def add_recurring_transaction():
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Name is required'}), 400
    
    try:
        start_date = datetime.fromisoformat(data['start_date'].replace('Z', ''))
        next_due_date = start_date
        end_date = datetime.fromisoformat(data['end_date'].replace('Z', '')) if data.get('end_date') and data['end_date'] else None
        
        if end_date and end_date < start_date:
            return jsonify({'error': 'End date cannot be before start date'}), 400

        today = datetime.utcnow().date()
        if start_date.date() < today:
             next_due_date = start_date
             while next_due_date.date() < today:
                 next_due = calculate_next_due(next_due_date, data['frequency'])
                 if not next_due: break
                 next_due_date = next_due

    except (ValueError, KeyError) as e:
        return jsonify({'error': f'Invalid date format or missing date: {e}'}), 400

    new_rt = RecurringTransaction(
        name=name,
        amount=float(data['amount']),
        frequency=data['frequency'],
        start_date=start_date,
        next_due_date=next_due_date,
        end_date=end_date,
        account_id=int(data['account_id']),
        category_id=int(data['category_id']) if data.get('category_id') else None,
        user_id=current_user.id
    )
    db.session.add(new_rt)
    db.session.commit()
    
    return jsonify({'id': new_rt.id}), 201

@app.route('/api/recurring/<int:rt_id>', methods=['PUT'])
@login_required
def update_recurring_transaction(rt_id):
    rt = RecurringTransaction.query.filter_by(id=rt_id, user_id=current_user.id).first_or_404()
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Name is required'}), 400

    try:
        start_date = datetime.fromisoformat(data['start_date'].replace('Z', ''))
        end_date = datetime.fromisoformat(data['end_date'].replace('Z', '')) if data.get('end_date') and data['end_date'] else None
        if end_date and end_date < start_date:
            return jsonify({'error': 'End date cannot be before start date'}), 400
        
        next_due_date = start_date
        today = datetime.utcnow().date()
        if next_due_date.date() < today:
             while next_due_date.date() < today:
                 next_due = calculate_next_due(next_due_date, data['frequency'])
                 if not next_due: break
                 next_due_date = next_due
        rt.next_due_date = next_due_date

    except (ValueError, KeyError) as e:
        return jsonify({'error': f'Invalid date format or missing date: {e}'}), 400

    rt.name = name
    rt.amount = float(data['amount'])
    rt.frequency = data['frequency']
    rt.start_date = start_date
    rt.end_date = end_date
    rt.account_id = int(data['account_id'])
    rt.category_id = int(data['category_id']) if data.get('category_id') else None
    
    db.session.commit()
    return jsonify({'message': 'Recurring transaction updated'})

@app.route('/api/recurring/<int:rt_id>', methods=['DELETE'])
@login_required
def delete_recurring_transaction(rt_id):
    rt = RecurringTransaction.query.filter_by(id=rt_id, user_id=current_user.id).first_or_404()
    db.session.delete(rt)
    db.session.commit()
    return jsonify({'message': 'Recurring transaction deleted'})

@app.route('/api/recurring/<int:rt_id>/toggle', methods=['PUT'])
@login_required
def toggle_recurring_transaction(rt_id):
    rt = RecurringTransaction.query.filter_by(id=rt_id, user_id=current_user.id).first_or_404()
    rt.is_active = not rt.is_active
    db.session.commit()
    return jsonify({'message': f'Recurring transaction set to {"active" if rt.is_active else "inactive"}'})

@app.route('/api/upload', methods=['POST'])
@login_required
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    account_id = request.form.get('account_id')
    use_ai = request.form.get('use_ai') == 'true'

    if not account_id:
        return jsonify({'error': 'Account ID is required'}), 400

    account = Account.query.filter_by(id=account_id, user_id=current_user.id).first()
    if not account:
        return jsonify({'error': 'Account not found'}), 404

    transactions_to_add = []
    descriptions_for_ai = []
    
    try:
        filename = file.filename.lower()
        if filename.endswith('.html') or filename.endswith('.htm'):
            # HTML Parsing Logic
            soup = BeautifulSoup(file.read(), 'html.parser')
            
            # Find the table with transaction headers
            tables = soup.find_all('table')
            target_table = None
            date_col_idx = -1
            desc_col_idx = -1
            amount_col_idx = -1
            
            for table in tables:
                headers = [th.get_text(strip=True) for th in table.find_all('th')]
                if not headers:
                    # check first row if no th present
                    first_row = table.find('tr')
                    if first_row:
                        headers = [td.get_text(strip=True) for td in first_row.find_all('td')]

                # Look for specific Turkish headers
                d_idx = next((i for i, h in enumerate(headers) if 'tarih' in h.lower()), -1)
                desc_idx = next((i for i, h in enumerate(headers) if 'açıklama' in h.lower()), -1)
                amt_idx = next((i for i, h in enumerate(headers) if 'tutar' in h.lower()), -1)
                
                if d_idx != -1 and desc_idx != -1 and amt_idx != -1:
                    target_table = table
                    date_col_idx = d_idx
                    desc_col_idx = desc_idx
                    amount_col_idx = amt_idx
                    break
            
            if not target_table:
                return jsonify({'error': 'Could not find transaction table in HTML.'}), 400

            # Iterate through rows, skipping header
            rows = target_table.find_all('tr')
            for row in rows:
                cells = row.find_all(['td', 'th'])
                if len(cells) <= max(date_col_idx, desc_col_idx, amount_col_idx):
                    continue
                
                # Verify it's a data row by checking if date cell looks like a date
                date_str = cells[date_col_idx].get_text(strip=True)
                try:
                    date_obj = datetime.strptime(date_str, '%d.%m.%Y')
                except ValueError:
                    continue # Skip header or invalid rows

                desc_str = cells[desc_col_idx].get_text(strip=True)
                amount_str = cells[amount_col_idx].get_text(strip=True)
                amount_val = parse_turkish_amount(amount_str)

                transactions_to_add.append({
                    'date': date_obj,
                    'description': desc_str,
                    'amount': amount_val
                })
                if desc_str:
                    descriptions_for_ai.append(desc_str)

        elif filename.endswith('.xlsx'):
            # Excel Parsing Logic (Keep existing logic mostly)
            wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            
            date_idx = next((i for i, h in enumerate(headers) if h and 'date' in str(h).lower()), None)
            desc_idx = next((i for i, h in enumerate(headers) if h and any(k in str(h).lower() for k in ['description', 'memo', 'payee', 'merchant', 'text', 'açıklama'])), None)
            amount_idx = next((i for i, h in enumerate(headers) if h and any(k in str(h).lower() for k in ['amount', 'betrag', 'value', 'price', 'tutar'])), None)

            if date_idx is None or desc_idx is None or amount_idx is None:
                 return jsonify({'error': 'Could not identify strict columns (Date, Description, Amount). Headers found: ' + str(headers)}), 400

            for row in ws.iter_rows(min_row=2, values_only=True):
                 if not row[date_idx] or row[amount_idx] is None: continue
                 
                 date_val = row[date_idx]
                 date_obj = datetime.utcnow()

                 if isinstance(date_val, str):
                     try:
                         date_obj = datetime.strptime(date_val, '%Y-%m-%d')
                     except: 
                         try:
                            date_obj = datetime.strptime(date_val, '%d.%m.%Y')
                         except:
                             continue 
                 elif isinstance(date_val, datetime):
                     date_obj = date_val
                 else:
                     continue

                 desc_val = str(row[desc_idx]).strip()
                 amount_val = row[amount_idx]
                 
                 if isinstance(amount_val, str):
                     try:
                        amount_val = float(re.sub(r'[^\d.-]', '', amount_val.replace(',', '.')))
                     except:
                        if ',' in amount_val and '.' in amount_val: # European/Turkish format using helper
                             amount_val = parse_turkish_amount(str(row[amount_idx]))
                        else:
                             continue

                 transactions_to_add.append({
                     'date': date_obj,
                     'description': desc_val,
                     'amount': float(amount_val)
                 })
                 if desc_val:
                    descriptions_for_ai.append(desc_val)
        else:
            return jsonify({'error': 'Unsupported file format. Please use .html or .xlsx'}), 400

        # --- AI and Saving Logic (Shared) ---
        category_map = {}
        processed_count = 0
        
        if use_ai and descriptions_for_ai:
            if not HAS_GENAI:
                logging.warning("AI categorization requested but google.generativeai module is missing.")
            else:
                api_key = current_user.gemini_api_key
                if api_key:
                    try:
                        genai.configure(api_key=api_key)
                        model = genai.GenerativeModel('gemini-2.5-flash')
                        
                        existing_categories = [c.name for c in Category.query.filter_by(user_id=current_user.id).all()]
                        unique_descriptions = list(set(descriptions_for_ai))[:50] 
                        
                        prompt = f"""
                        You are a financial assistant. Map these transaction descriptions to the most appropriate category from this list: {existing_categories}.
                        If none fit well, create a new simple category name (e.g. 'Groceries', 'Transport', 'Utilities', 'Salary', 'Transfer').
                        Respond ONLY with a valid JSON object where keys are descriptions and values are category names.
                        Descriptions: {json.dumps(unique_descriptions)}
                        """
                        
                        response = model.generate_content(prompt)
                        text_response = response.text.strip()
                        if text_response.startswith('```json'):
                            text_response = text_response[7:-3]
                        elif text_response.startswith('```'): # Handle plain code block
                            text_response = text_response[3:-3]
                        
                        category_map = json.loads(text_response)
                    except Exception as ai_error:
                        logging.error(f"AI Categorization failed: {ai_error}")

        for tx in transactions_to_add:
            # Check for dupes (simple check based on date, amount, description)
            # This is optional but good practice to avoid re-run spam
            exists = Transaction.query.filter_by(
                account_id=account.id, 
                date=tx['date'], 
                amount=tx['amount'], 
                description=tx['description']
            ).first()
            
            if exists:
                continue

            category_id = None
            cat_name = category_map.get(tx['description'])
            
            if cat_name:
                category = Category.query.filter_by(user_id=current_user.id, name=cat_name).first()
                if not category:
                    # Determine type based on amount sign usually, but let's default to expense unless positive
                    cat_type = 'income' if tx['amount'] > 0 else 'expense'
                    category = Category(name=cat_name, type=cat_type, user_id=current_user.id)
                    db.session.add(category)
                    db.session.commit()
                category_id = category.id

            new_trans = Transaction(
                description=tx['description'],
                amount=tx['amount'],
                date=tx['date'],
                account_id=account.id,
                category_id=category_id
            )
            db.session.add(new_trans)
            account.balance += tx['amount']
            processed_count += 1

        db.session.commit()
        return jsonify({'message': 'Import successful', 'count': processed_count})

    except Exception as e:
        logging.error(f"Import failed: {e}")
        return jsonify({'error': f"Import failed: {str(e)}"}), 500